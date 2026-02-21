"""Attachment content extraction — PDF, Excel, CSV, images."""

from __future__ import annotations

import csv
import io

from tunadex.extraction.schema import AttachmentType


def extract_text_from_pdf(pdf_bytes: bytes) -> str:
    """Extract text from PDF using pdfplumber."""
    import pdfplumber

    text_parts: list[str] = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)

            # Also try extracting tables
            tables = page.extract_tables()
            for table in tables:
                for row in table:
                    cells = [str(c) if c else "" for c in row]
                    text_parts.append(" | ".join(cells))

    return "\n".join(text_parts)


def extract_rows_from_excel(excel_bytes: bytes) -> list[dict]:
    """Extract rows from Excel file as list of dicts."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(excel_bytes), read_only=True, data_only=True)
    all_rows: list[dict] = []

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            continue

        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        for row in rows[1:]:
            row_dict = {}
            for i, val in enumerate(row):
                key = headers[i] if i < len(headers) else f"col_{i}"
                row_dict[key] = val
            if any(v is not None for v in row_dict.values()):
                all_rows.append(row_dict)

    wb.close()
    return all_rows


def extract_text_from_csv(csv_bytes: bytes) -> str:
    """Extract text from CSV file."""
    text = csv_bytes.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    lines: list[str] = []
    for row in reader:
        lines.append(" | ".join(row))
    return "\n".join(lines)


def extract_text_from_attachment(
    file_bytes: bytes, attachment_type: AttachmentType
) -> str:
    """Extract readable text from an attachment based on its type.

    For images, returns empty string — use Gemini multimodal instead.
    """
    if attachment_type == AttachmentType.PDF:
        return extract_text_from_pdf(file_bytes)
    elif attachment_type == AttachmentType.EXCEL:
        rows = extract_rows_from_excel(file_bytes)
        lines = []
        for row in rows:
            parts = [f"{k}: {v}" for k, v in row.items() if v is not None]
            lines.append(" | ".join(parts))
        return "\n".join(lines)
    elif attachment_type == AttachmentType.CSV:
        return extract_text_from_csv(file_bytes)
    elif attachment_type == AttachmentType.IMAGE:
        return ""  # Handled by Gemini multimodal
    else:
        try:
            return file_bytes.decode("utf-8", errors="replace")[:5000]
        except Exception:
            return ""
